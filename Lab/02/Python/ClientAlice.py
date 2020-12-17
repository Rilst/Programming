from flask import Flask, request
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font

Buf_Size = 1
buf = []

def FindClearRow(page, col):
    i = 1
    while True:
        if type(page[i][col]).__name__ == 'MergedCell':
            i += 1
        else:
            if page[i][col].value != None:
                i += 1
            else:
                break
    return i

def writeinbuf(Res):
    global buf
    global Buf_Size
    buf.append(Res);
    buf[-1]["datetime"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    if len(buf) >= Buf_Size:
        try:
            book = openpyxl.open("data.xlsx", read_only=False)
        except:
            book = openpyxl.Workbook()
        page = book.active
        if page.title != "Alice":
        
            page.title = "Alice"
            page['A1'] = "N"
            page['B1'] = "User ID"
            page['C1'] = "Datetime"
            page['D1'] = "Item"
            page['E1'] = "Price"
            
            page.column_dimensions['A'].width = 5
            page.column_dimensions['B'].width = 67
            page.column_dimensions['C'].width = 18
            page.column_dimensions['D'].width = 30
            page.column_dimensions['E'].width = 13
            
            page[1][0].alignment = Alignment(horizontal='center', vertical='center')
            page[1][1].alignment = Alignment(horizontal='center', vertical='center')
            page[1][2].alignment = Alignment(horizontal='center', vertical='center')
            page[1][3].alignment = Alignment(horizontal='center', vertical='center')
            page[1][4].alignment = Alignment(horizontal='center', vertical='center')
            page[1][0].font = Font(bold=True)
            page[1][1].font = Font(bold=True)
            page[1][2].font = Font(bold=True)
            page[1][3].font = Font(bold=True)
            page[1][4].font = Font(bold=True)
            
        for i in range(len(buf)):
        
            N = FindClearRow(page, 0)
            OLD = N
            while True:
                N -= 1
                if type(page[N][0]).__name__ == 'Cell':
                    if str(page[N][0].value).isdigit():
                        page[OLD][0].value = page[N][0].value + 1
                        page[OLD][0].alignment = Alignment(horizontal='center', vertical='center')
                        break
                    else:
                        page[OLD][0].value = 1
                        page[OLD][0].alignment = Alignment(horizontal='center', vertical='center')
                        break
                
            UID = FindClearRow(page, 1)
            page[UID][1].value = buf[i]["user_id"]
            page[UID][1].alignment = Alignment(horizontal='center', vertical='center')
            
            DateT = FindClearRow(page, 2)
            page[DateT][2].value = buf[i]["datetime"]
            page[DateT][2].alignment = Alignment(horizontal='center', vertical='center')
            
            Leng = len(buf[i]["check"])
            page.merge_cells(start_row=OLD, start_column=1, end_row=OLD+Leng-1, end_column=1)
            page.merge_cells(start_row=OLD, start_column=2, end_row=OLD+Leng-1, end_column=2)
            page.merge_cells(start_row=OLD, start_column=3, end_row=OLD+Leng-1, end_column=3)
            
            for K in range(len(buf[i]["check"])):
                Itm = FindClearRow(page, 3)
                page[Itm][3].value = buf[i]["check"][K]["item"]
                page[Itm][3].alignment = Alignment(horizontal='center', vertical='center')
                page[Itm][4].value = buf[i]["check"][K]["price"]
                page[Itm][4].alignment = Alignment(horizontal='center', vertical='center')    
        book.save("data.xlsx")
        buf = []

app = Flask(__name__)

@app.route('/', methods=['POST'])
def index():
    if request.method == 'POST':
        if request.is_json:
            Res = request.get_json()
            writeinbuf(Res)
            return "OK"
 
if __name__ == "__main__":
    app.run()